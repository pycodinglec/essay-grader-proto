"""리포트 생성 모듈 (report.xlsx).

채점 결과를 취합하여 11개 컬럼(학번/이름/작품번호/에세이/모델별 점수·피드백/최종 점수)을
가진 report.xlsx 바이트를 생성한다.
"""

from __future__ import annotations

from collections import defaultdict
from io import BytesIO
from typing import Optional

import openpyxl

from src.evaluator import sum_scores

_HEADERS = [
    "학번", "이름", "작품번호", "에세이(Gemini-3-pro-preview)",
    "합산 점수(GPT)", "합산 점수(Gemini)", "합산 점수(Claude)",
    "피드백(GPT)", "피드백(Gemini)", "피드백(Claude)", "최종 점수",
]


def _model_total_score(model_eval: Optional[dict]) -> int | float | str:
    """모델 평가 결과에서 점수 합산을 반환한다. 실패 시 빈 문자열."""
    if model_eval is None:
        return ""
    return sum(item["점수"] for item in model_eval["scores"])


def _model_feedback(model_eval: Optional[dict]) -> str:
    """모델 평가 결과에서 피드백을 반환한다. 실패 시 빈 문자열."""
    if model_eval is None:
        return ""
    return model_eval["feedback"]


def assign_work_numbers(submissions: list[dict]) -> list[dict]:
    """제출물 리스트에 학생별 작품번호를 부여한다.

    동일 학번의 제출물은 등장 순서에 따라 1, 2, 3... 번호를 받는다.
    평가결과에서 모델별 점수/피드백과 최종 점수를 추출한다.

    Args:
        submissions: {"학번", "이름", "에세이텍스트",
                      "평가결과": {"best", "by_model"}} 리스트.

    Returns:
        11개 키를 가진 dict 리스트.
    """
    counters: dict[str, int] = defaultdict(int)
    result: list[dict] = []

    for sub in submissions:
        student_id = sub["학번"]
        counters[student_id] += 1
        evaluation = sub["평가결과"]
        by_model = evaluation["by_model"]
        best = evaluation["best"]
        result.append({
            "학번": student_id,
            "이름": sub["이름"],
            "작품번호": counters[student_id],
            "에세이(Gemini-3-pro-preview)": sub["에세이텍스트"],
            "합산 점수(GPT)": _model_total_score(by_model.get("openai")),
            "합산 점수(Gemini)": _model_total_score(by_model.get("gemini")),
            "합산 점수(Claude)": _model_total_score(by_model.get("anthropic")),
            "피드백(GPT)": _model_feedback(by_model.get("openai")),
            "피드백(Gemini)": _model_feedback(by_model.get("gemini")),
            "피드백(Claude)": _model_feedback(by_model.get("anthropic")),
            "최종 점수": sum_scores(best),
        })

    return result


def generate_report_bytes(report_data: list[dict]) -> bytes:
    """리포트 데이터를 xlsx 바이트로 변환한다.

    행은 학번 오름차순, 그 다음 작품번호 오름차순으로 정렬된다.

    Args:
        report_data: 11개 키를 가진 dict 리스트.

    Returns:
        xlsx 파일의 바이트 데이터.
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    for col_idx, header in enumerate(_HEADERS, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    sorted_data = sorted(
        report_data, key=lambda r: (r["학번"], r["작품번호"])
    )

    for row_idx, record in enumerate(sorted_data, start=2):
        for col_idx, key in enumerate(_HEADERS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=record[key])

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_report(graded_submissions: list[dict]) -> bytes:
    """채점 완료된 제출물로부터 report.xlsx 바이트를 생성한다.

    assign_work_numbers와 generate_report_bytes를 순차 호출하는
    상위 수준 함수이다.

    Args:
        graded_submissions: {"학번", "이름", "에세이텍스트",
                             "평가결과": {"best", "by_model"}} dict 리스트.

    Returns:
        다운로드 가능한 xlsx 바이트.
    """
    report_data = assign_work_numbers(graded_submissions)
    return generate_report_bytes(report_data)
