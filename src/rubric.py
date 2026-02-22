"""채점기준표 검증 모듈.

xlsx 파일에서 채점기준표를 읽고, 양식을 검증하고,
파싱된 데이터를 반환하는 함수를 제공한다.
"""

from io import BytesIO

from openpyxl import load_workbook

_EXPECTED_HEADERS = ["번호", "채점기준", "배점"]


def validate_rubric(file_bytes: bytes) -> tuple[bool, str]:
    """xlsx 바이트를 받아 채점기준표 양식을 검증한다.

    Args:
        file_bytes: xlsx 파일의 바이트 데이터.

    Returns:
        (True, "") 유효한 경우, (False, 에러메시지) 유효하지 않은 경우.
    """
    try:
        wb = load_workbook(filename=BytesIO(file_bytes), read_only=True)
    except Exception:
        return False, "유효한 xlsx 파일이 아닙니다."

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return False, "파일에 데이터가 없습니다."

    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]

    if headers != _EXPECTED_HEADERS:
        return False, (
            f"헤더가 올바르지 않습니다. "
            f"필요: {_EXPECTED_HEADERS}, 실제: {headers}"
        )

    data_rows = rows[1:]
    if not data_rows:
        return False, "데이터 행이 최소 1개 이상 필요합니다."

    for i, row in enumerate(data_rows, start=2):
        score = row[2] if len(row) > 2 else None
        if not isinstance(score, (int, float)):
            return False, f"{i}행의 배점 값이 숫자가 아닙니다: {score}"

    return True, ""


def parse_rubric(file_bytes: bytes) -> list[dict]:
    """xlsx 바이트를 받아 채점기준표 데이터를 파싱한다.

    Args:
        file_bytes: xlsx 파일의 바이트 데이터.

    Returns:
        [{"번호": int, "채점기준": str, "배점": int|float}, ...] 형태의 리스트.

    Raises:
        ValueError: 파일이 유효하지 않을 경우.
    """
    ok, msg = validate_rubric(file_bytes)
    if not ok:
        raise ValueError(msg)

    wb = load_workbook(filename=BytesIO(file_bytes), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    result = []
    for row in rows[1:]:
        result.append({
            "번호": row[0],
            "채점기준": str(row[1]),
            "배점": row[2],
        })

    return result


def format_rubric_for_display(rubric_data: list[dict]) -> str:
    """채점기준표 데이터를 사용자에게 보여줄 문자열로 변환한다.

    Args:
        rubric_data: parse_rubric이 반환하는 dict 리스트.

    Returns:
        "번호 | 채점기준 | 배점" 형식의 문자열.
    """
    lines = ["번호 | 채점기준 | 배점"]

    for item in rubric_data:
        lines.append(
            f"{item['번호']} | {item['채점기준']} | {item['배점']}"
        )

    return "\n".join(lines)
