"""report 모듈 단위 테스트."""

from io import BytesIO

import openpyxl

from src.report import assign_work_numbers, build_report, generate_report_bytes

_HEADERS = [
    "학번", "이름", "작품번호", "에세이(Gemini-3-pro-preview)",
    "합산 점수(GPT)", "합산 점수(Gemini)", "합산 점수(Claude)",
    "피드백(GPT)", "피드백(Gemini)", "피드백(Claude)", "최종 점수",
]
_COL_COUNT = len(_HEADERS)


def _make_submission(
    student_id: str,
    name: str,
    essay_text: str = "에세이 본문",
    gemini_score: int = 5,
    openai_score: int = 7,
    anthropic_score: int = 6,
    gemini_feedback: str = "G피드백",
    openai_feedback: str = "O피드백",
    anthropic_feedback: str = "A피드백",
    failed_models: tuple[str, ...] = (),
) -> dict:
    """테스트용 제출물 dict를 생성하는 헬퍼."""
    by_model: dict = {}
    best = None
    best_score = -1

    for model, score, fb in [
        ("gemini", gemini_score, gemini_feedback),
        ("openai", openai_score, openai_feedback),
        ("anthropic", anthropic_score, anthropic_feedback),
    ]:
        if model in failed_models:
            by_model[model] = None
        else:
            entry = {"scores": [{"번호": 1, "점수": score}], "feedback": fb}
            by_model[model] = entry
            if score > best_score:
                best_score = score
                best = entry

    return {
        "학번": student_id,
        "이름": name,
        "에세이텍스트": essay_text,
        "평가결과": {"best": best, "by_model": by_model},
    }


def _make_report_row(
    student_id: str = "10305",
    name: str = "홍길동",
    work_num: int = 1,
    essay: str = "에세이 본문",
    gpt_score: int | str = 7,
    gemini_score: int | str = 5,
    claude_score: int | str = 6,
    gpt_feedback: str = "O피드백",
    gemini_feedback: str = "G피드백",
    claude_feedback: str = "A피드백",
    final_score: int | str = 7,
) -> dict:
    """테스트용 리포트 데이터 dict를 생성하는 헬퍼."""
    return {
        "학번": student_id,
        "이름": name,
        "작품번호": work_num,
        "에세이(Gemini-3-pro-preview)": essay,
        "합산 점수(GPT)": gpt_score,
        "합산 점수(Gemini)": gemini_score,
        "합산 점수(Claude)": claude_score,
        "피드백(GPT)": gpt_feedback,
        "피드백(Gemini)": gemini_feedback,
        "피드백(Claude)": claude_feedback,
        "최종 점수": final_score,
    }


class TestAssignWorkNumbers:
    """assign_work_numbers 함수 테스트."""

    def test_single_student_single_work(self):
        """학생 1명, 작품 1개일 때 작품번호 1이 부여된다."""
        submissions = [_make_submission("10305", "홍길동")]

        result = assign_work_numbers(submissions)

        assert len(result) == 1
        assert result[0]["작품번호"] == 1

    def test_single_student_multiple_works(self):
        """학생 1명, 작품 여러 개일 때 순서대로 1, 2, 3이 부여된다."""
        submissions = [
            _make_submission("10305", "홍길동"),
            _make_submission("10305", "홍길동"),
            _make_submission("10305", "홍길동"),
        ]

        result = assign_work_numbers(submissions)

        assert [r["작품번호"] for r in result] == [1, 2, 3]

    def test_multiple_students_independent_numbering(self):
        """서로 다른 학생은 각각 독립적으로 작품번호가 부여된다."""
        submissions = [
            _make_submission("10305", "홍길동"),
            _make_submission("20101", "김철수"),
            _make_submission("10305", "홍길동"),
            _make_submission("20101", "김철수"),
        ]

        result = assign_work_numbers(submissions)

        hong = [r for r in result if r["학번"] == "10305"]
        kim = [r for r in result if r["학번"] == "20101"]
        assert [h["작품번호"] for h in hong] == [1, 2]
        assert [k["작품번호"] for k in kim] == [1, 2]

    def test_output_has_11_keys(self):
        """반환되는 dict에 11개 키가 존재한다."""
        submissions = [_make_submission("10305", "홍길동")]

        result = assign_work_numbers(submissions)

        assert set(result[0].keys()) == set(_HEADERS)

    def test_model_scores_extracted(self):
        """모델별 점수가 올바르게 추출된다."""
        submissions = [_make_submission(
            "10305", "홍길동",
            gemini_score=8, openai_score=10, anthropic_score=9,
        )]

        result = assign_work_numbers(submissions)

        assert result[0]["합산 점수(Gemini)"] == 8
        assert result[0]["합산 점수(GPT)"] == 10
        assert result[0]["합산 점수(Claude)"] == 9

    def test_model_feedbacks_extracted(self):
        """모델별 피드백이 올바르게 추출된다."""
        submissions = [_make_submission(
            "10305", "홍길동",
            gemini_feedback="G좋음", openai_feedback="O우수",
            anthropic_feedback="A양호",
        )]

        result = assign_work_numbers(submissions)

        assert result[0]["피드백(Gemini)"] == "G좋음"
        assert result[0]["피드백(GPT)"] == "O우수"
        assert result[0]["피드백(Claude)"] == "A양호"

    def test_failed_model_shows_empty_string(self):
        """실패 모델은 빈 문자열로 표시된다."""
        submissions = [_make_submission(
            "10305", "홍길동", failed_models=("gemini", "anthropic"),
        )]

        result = assign_work_numbers(submissions)

        assert result[0]["합산 점수(Gemini)"] == ""
        assert result[0]["피드백(Gemini)"] == ""
        assert result[0]["합산 점수(Claude)"] == ""
        assert result[0]["피드백(Claude)"] == ""
        assert result[0]["합산 점수(GPT)"] == 7
        assert result[0]["피드백(GPT)"] == "O피드백"

    def test_final_score_from_best(self):
        """최종 점수는 best의 합산 점수이다."""
        submissions = [_make_submission(
            "10305", "홍길동",
            gemini_score=5, openai_score=10, anthropic_score=8,
        )]

        result = assign_work_numbers(submissions)

        assert result[0]["최종 점수"] == 10

    def test_essay_text_preserved(self):
        """에세이 원문이 보존된다."""
        submissions = [_make_submission(
            "10305", "홍길동", essay_text="원본 에세이 텍스트",
        )]

        result = assign_work_numbers(submissions)

        assert result[0]["에세이(Gemini-3-pro-preview)"] == "원본 에세이 텍스트"

    def test_empty_input(self):
        """빈 리스트 입력 시 빈 리스트를 반환한다."""
        result = assign_work_numbers([])

        assert result == []

    def test_student_name_preserved(self):
        """학번과 이름이 원본 값 그대로 유지된다."""
        submissions = [_make_submission("30210", "이영희")]

        result = assign_work_numbers(submissions)

        assert result[0]["학번"] == "30210"
        assert result[0]["이름"] == "이영희"


class TestGenerateReportBytes:
    """generate_report_bytes 함수 테스트."""

    def _load_workbook(self, xlsx_bytes: bytes) -> openpyxl.Workbook:
        """바이트에서 워크북을 로드하는 헬퍼."""
        return openpyxl.load_workbook(BytesIO(xlsx_bytes))

    def test_returns_bytes(self):
        """반환값이 bytes 타입이다."""
        result = generate_report_bytes([_make_report_row()])

        assert isinstance(result, bytes)

    def test_valid_xlsx(self):
        """반환된 bytes가 유효한 xlsx 파일이다."""
        result = generate_report_bytes([_make_report_row()])
        wb = self._load_workbook(result)

        assert wb.active is not None

    def test_correct_headers(self):
        """첫 행에 11개 헤더가 있다."""
        result = generate_report_bytes([_make_report_row()])
        ws = self._load_workbook(result).active

        headers = [ws.cell(row=1, column=c).value for c in range(1, _COL_COUNT + 1)]
        assert headers == _HEADERS

    def test_correct_data_rows(self):
        """데이터 행이 올바르게 기록된다."""
        data = [_make_report_row(
            gpt_score=10, gemini_score=8, claude_score=9,
            gpt_feedback="GPT좋음", gemini_feedback="Gem좋음",
            claude_feedback="Claude좋음", final_score=10,
            essay="에세이 원문",
        )]

        result = generate_report_bytes(data)
        ws = self._load_workbook(result).active

        row = [ws.cell(row=2, column=c).value for c in range(1, _COL_COUNT + 1)]
        assert row == [
            "10305", "홍길동", 1, "에세이 원문",
            10, 8, 9,
            "GPT좋음", "Gem좋음", "Claude좋음", 10,
        ]

    def test_sorted_by_student_id_then_work_number(self):
        """행이 학번 오름차순, 그 다음 작품번호 오름차순으로 정렬된다."""
        data = [
            _make_report_row(student_id="20101", name="김철수", work_num=2),
            _make_report_row(student_id="10305", work_num=1),
            _make_report_row(student_id="20101", name="김철수", work_num=1),
            _make_report_row(student_id="10305", work_num=2),
        ]

        result = generate_report_bytes(data)
        ws = self._load_workbook(result).active

        rows = []
        for r in range(2, 6):
            rows.append([ws.cell(row=r, column=c).value for c in range(1, _COL_COUNT + 1)])

        assert rows[0][0] == "10305"
        assert rows[0][2] == 1
        assert rows[1][0] == "10305"
        assert rows[1][2] == 2
        assert rows[2][0] == "20101"
        assert rows[2][2] == 1
        assert rows[3][0] == "20101"
        assert rows[3][2] == 2

    def test_empty_data_returns_headers_only(self):
        """빈 데이터 시 헤더만 있는 xlsx를 반환한다."""
        result = generate_report_bytes([])
        ws = self._load_workbook(result).active

        headers = [ws.cell(row=1, column=c).value for c in range(1, _COL_COUNT + 1)]
        assert headers == _HEADERS
        assert ws.cell(row=2, column=1).value is None

    def test_multiple_rows_count(self):
        """데이터 행 수가 입력과 일치한다 (헤더 제외)."""
        data = [
            _make_report_row(work_num=1),
            _make_report_row(work_num=2),
            _make_report_row(student_id="20101", name="김철수", work_num=1),
        ]

        result = generate_report_bytes(data)
        ws = self._load_workbook(result).active

        assert ws.max_row == 4  # 1 header + 3 data rows


class TestBuildReport:
    """build_report 함수 통합 테스트."""

    def test_returns_bytes(self):
        """build_report는 bytes를 반환한다."""
        submissions = [_make_submission("10305", "홍길동")]

        result = build_report(submissions)

        assert isinstance(result, bytes)

    def test_end_to_end_single_student(self):
        """단일 학생 제출물에 대해 올바른 xlsx를 생성한다."""
        submissions = [_make_submission(
            "10305", "홍길동", essay_text="원본 텍스트",
            openai_score=10, gemini_score=8, anthropic_score=9,
        )]

        result = build_report(submissions)
        ws = openpyxl.load_workbook(BytesIO(result)).active

        headers = [ws.cell(row=1, column=c).value for c in range(1, _COL_COUNT + 1)]
        assert headers == _HEADERS

        assert ws.cell(row=2, column=1).value == "10305"
        assert ws.cell(row=2, column=4).value == "원본 텍스트"
        assert ws.cell(row=2, column=5).value == 10   # GPT
        assert ws.cell(row=2, column=6).value == 8    # Gemini
        assert ws.cell(row=2, column=7).value == 9    # Claude
        assert ws.cell(row=2, column=11).value == 10  # 최종 점수

    def test_end_to_end_multiple_students_sorted(self):
        """다수 학생 제출물이 학번/작품번호 기준 정렬된 xlsx를 생성한다."""
        submissions = [
            _make_submission("20101", "김철수"),
            _make_submission("10305", "홍길동"),
            _make_submission("20101", "김철수"),
        ]

        result = build_report(submissions)
        ws = openpyxl.load_workbook(BytesIO(result)).active

        assert ws.cell(row=2, column=1).value == "10305"
        assert ws.cell(row=2, column=3).value == 1
        assert ws.cell(row=3, column=1).value == "20101"
        assert ws.cell(row=3, column=3).value == 1
        assert ws.cell(row=4, column=1).value == "20101"
        assert ws.cell(row=4, column=3).value == 2

    def test_empty_submissions(self):
        """빈 제출물 리스트에 대해 헤더만 있는 xlsx를 반환한다."""
        result = build_report([])
        ws = openpyxl.load_workbook(BytesIO(result)).active

        headers = [ws.cell(row=1, column=c).value for c in range(1, _COL_COUNT + 1)]
        assert headers == _HEADERS
        assert ws.cell(row=2, column=1).value is None
